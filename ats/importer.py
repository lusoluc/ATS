"""P0.5: CSV-Bewerberdaten-Import – die minimale Migrationsbrücke (ROADMAP V0).

Design-Entscheidungen (bewusst konservativ, es sind Bestands-PII):
- **Testlauf zuerst:** `dry_run=True` liefert denselben Bericht, ändert aber nichts.
  Die UI bietet „Prüfen" als eigenen Button – Vertrauen vor Tempo.
- **Keine Duplikate:** Bewerber werden über den E-Mail-Blind-Index wiedererkannt
  (`get_or_create_by_email`); eine bereits vorhandene Bewerbung (Bewerber+Stelle)
  wird übersprungen, nicht dupliziert.
- **Keine erfundenen Stellen:** Die Spalte „Stelle" muss eine existierende
  Ausschreibung treffen (Titel, case-insensitive) – sonst Fehlerzeile. Optional
  kann eine Standard-Stelle für Zeilen ohne Angabe gewählt werden.
- **Excel-tolerant:** UTF-8 mit/ohne BOM, Semikolon ODER Komma (Auto-Erkennung),
  deutsche und englische Spaltenköpfe.
"""
import csv
import io
from datetime import datetime

from django.utils import timezone

MAX_ROWS = 5000

# Spaltenkopf-Aliasse (normalisiert: lower, ohne Leerzeichen/Umlaut-Punkte)
HEADER_ALIASES = {
    "first_name": {"vorname", "firstname", "first_name", "givenname"},
    "last_name": {"nachname", "name", "lastname", "last_name", "surname", "familienname"},
    "email": {"email", "e-mail", "mail", "emailadresse", "e-mail-adresse"},
    "phone": {"telefon", "phone", "tel", "telefonnummer", "mobil", "handy"},
    "job": {"stelle", "job", "position", "stellenanzeige", "ausschreibung", "jobtitel"},
    "status": {"status"},
    "source": {"quelle", "source", "kanal"},
    "address": {"adresse", "anschrift", "strasse", "straße", "street",
                "address", "wohnort"},
    "notes": {"notiz", "notizen", "notes", "anmerkung", "kommentar"},
    "created": {"datum", "eingegangen", "created", "bewerbungsdatum", "date"},
}

STATUS_ALIASES = {
    "new": "NEW", "neu": "NEW", "": "NEW",
    "in_review": "IN_REVIEW", "in prüfung": "IN_REVIEW", "in pruefung": "IN_REVIEW",
    "prüfung": "IN_REVIEW", "review": "IN_REVIEW",
    "invited": "INVITED", "eingeladen": "INVITED", "interview": "INVITED",
    "rejected": "REJECTED", "abgelehnt": "REJECTED", "absage": "REJECTED",
    "withdrawn": "WITHDRAWN", "zurückgezogen": "WITHDRAWN", "zurueckgezogen": "WITHDRAWN",
    "missing_docs": "MISSING_DOCS", "unterlagen fehlen": "MISSING_DOCS",
}


def _norm_header(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "").replace("_", "").replace("-", "")


def _map_headers(fieldnames, overrides=None):
    """Ordnet CSV-Spalten unseren Feldern zu; unbekannte Spalten werden
    ignoriert. `overrides` ({feld: spaltenname}) kommt aus der manuellen
    Zuordnung im Import-Dialog und GEWINNT gegen die Synonym-Erkennung –
    aber nur fuer Spalten, die es wirklich gibt (nichts Erfundenes)."""
    mapping = {}
    normalized_aliases = {
        field: {_norm_header(a) for a in aliases}
        for field, aliases in HEADER_ALIASES.items()
    }
    for raw in fieldnames or []:
        n = _norm_header(raw)
        for field, aliases in normalized_aliases.items():
            if n in aliases and field not in mapping:
                mapping[field] = raw
    for field, col in (overrides or {}).items():
        if field in HEADER_ALIASES and col in (fieldnames or []):
            mapping[field] = col
        elif field in HEADER_ALIASES and col == "__IGNORE__":
            mapping.pop(field, None)
    return mapping


def detect_headers(file_bytes, is_xlsx=False):
    """Nur die Kopfzeile lesen – fuer den Zuordnungs-Dialog im Import."""
    if is_xlsx:
        import io as _io
        try:
            from openpyxl import load_workbook
            wb = load_workbook(_io.BytesIO(file_bytes), read_only=True)
            header = next(wb.worksheets[0].iter_rows(values_only=True), ())
            wb.close()
            return [str(h).strip() for h in header if h is not None]
        except Exception:
            return []
    import csv as _csv
    import io as _io
    try:
        text = file_bytes.decode("utf-8-sig", errors="replace")
        sample = text[:2048]
        try:
            dialect = _csv.Sniffer().sniff(sample, delimiters=";,")
        except _csv.Error:
            class dialect(_csv.excel):
                delimiter = ";"
        reader = _csv.reader(_io.StringIO(text), dialect=dialect)
        return [h.strip() for h in next(reader, []) if h.strip()]
    except Exception:
        return []


def _parse_date(value):
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y"):
        try:
            dt = datetime.strptime(value.strip(), fmt)
            return timezone.make_aware(dt) if timezone.is_naive(dt) else dt
        except (ValueError, AttributeError):
            continue
    return None


def parse_csv(file_bytes: bytes, overrides=None):
    """Bytes → (rows, fatal_error). rows = Liste normalisierter Dicts mit _line."""
    try:
        text = file_bytes.decode("utf-8-sig")  # toleriert Excel-BOM
    except UnicodeDecodeError:
        try:
            text = file_bytes.decode("latin-1")  # Alt-Excel-Fallback
        except UnicodeDecodeError:
            return [], "Datei ist weder UTF-8 noch Latin-1 kodiert."

    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,")
    except csv.Error:
        class dialect:  # noqa: N801 – csv erwartet Klasse/Instanz
            delimiter = ";"
            quotechar = '"'

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    mapping = _map_headers(reader.fieldnames, overrides)
    missing = [f for f in ("first_name", "last_name", "email") if f not in mapping]
    if missing:
        return [], ("Pflichtspalten fehlen: " + ", ".join(missing) +
                    ". Erwartet werden mindestens Vorname, Nachname, E-Mail "
                    "(deutsche oder englische Spaltenköpfe).")

    rows = []
    for i, raw in enumerate(reader, start=2):  # Zeile 1 = Kopf
        if i - 1 > MAX_ROWS:
            return [], f"Mehr als {MAX_ROWS} Zeilen – bitte Datei aufteilen."
        row = {field: (raw.get(col) or "").strip()
               for field, col in mapping.items()}
        row["_line"] = i
        if not any(v for k, v in row.items() if k != "_line"):
            continue  # komplett leere Zeile still überspringen
        rows.append(row)
    return rows, None


def run_import(rows, default_job=None, dry_run=True):
    """Importiert (oder simuliert) die Zeilen. Gibt einen ehrlichen Bericht zurück."""
    from django.db import transaction

    from .models import Applicant, Application, JobPosting, email_blind_index

    jobs_by_title = {j.title.strip().lower(): j for j in JobPosting.objects.all()}
    report = {
        "total": len(rows),
        "applicants_new": 0, "applicants_reused": 0,
        "applications_created": 0, "skipped_existing": 0,
        "errors": [],  # (Zeile, Grund)
        "dry_run": dry_run,
    }
    seen_pairs = set()  # (emailhash-proxy, job_id) innerhalb der Datei

    with transaction.atomic():
        for row in rows:
            line = row["_line"]
            email = row.get("email", "").lower()
            if not row.get("first_name") or not row.get("last_name"):
                report["errors"].append((line, "Vor- oder Nachname fehlt"))
                continue
            if "@" not in email or "." not in email.rsplit("@", 1)[-1]:
                report["errors"].append((line, f"Ungültige E-Mail: '{row.get('email')}'"))
                continue

            job = default_job
            if row.get("job"):
                job = jobs_by_title.get(row["job"].lower())
                if job is None:
                    report["errors"].append(
                        (line, f"Stelle nicht gefunden: '{row['job']}' – Import legt "
                               "bewusst keine Ausschreibungen an"))
                    continue
            if job is None:
                report["errors"].append(
                    (line, "Keine Stelle angegeben und keine Standard-Stelle gewählt"))
                continue

            status = STATUS_ALIASES.get(row.get("status", "").strip().lower())
            if status is None:
                report["errors"].append(
                    (line, f"Unbekannter Status: '{row['status']}' (erlaubt: neu, "
                           "in Prüfung, eingeladen, abgelehnt, zurückgezogen)"))
                continue

            pair_key = (email, str(job.id))
            existing = Applicant.objects.filter(
                emailHash=email_blind_index(email)).first()

            if (existing and Application.objects.filter(
                    applicant=existing, jobPosting=job).exists()) or pair_key in seen_pairs:
                report["skipped_existing"] += 1
                continue
            seen_pairs.add(pair_key)

            if dry_run:
                if existing:
                    report["applicants_reused"] += 1
                else:
                    report["applicants_new"] += 1
                report["applications_created"] += 1
                continue

            applicant, created = Applicant.objects.get_or_create_by_email(
                email, defaults={"firstName": row["first_name"][:100],
                                 "lastName": row["last_name"][:100],
                                 "phone": row.get("phone", "")[:50] or None,
                                 "address": row.get("address", "")[:300] or None})
            report["applicants_new" if created else "applicants_reused"] += 1

            app = Application.objects.create(
                applicant=applicant, jobPosting=job, status=status,
                source=(row.get("source") or "IMPORT").upper()[:50],
                internalNotes=(row.get("notes") or "")[:2000],
            )
            created_at = _parse_date(row.get("created", ""))
            if created_at:
                Application.objects.filter(id=app.id).update(createdAt=created_at)
            report["applications_created"] += 1

        if dry_run:
            transaction.set_rollback(True)  # Testlauf: garantiert keine Änderung
    return report


def parse_xlsx(file_bytes: bytes, overrides=None):
    """XLSX → (rows, fatal_error) im selben Format wie parse_csv.

    Bestandssysteme exportieren fast immer Excel – der Umstieg darf nicht an
    'bitte erst als CSV speichern' scheitern. Erste Tabelle, erste Zeile =
    Kopfzeile, gleiches Spalten-Mapping wie CSV (dieselben Synonyme).
    """
    import io as _io
    try:
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(file_bytes), read_only=True,
                           data_only=True)
    except Exception:
        return [], ("Datei ist kein lesbares XLSX. Bitte als .xlsx exportieren "
                    "oder den CSV-Weg nutzen.")
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return [], "Die Tabelle ist leer."
    fieldnames = [str(h).strip() if h is not None else "" for h in header]
    mapping = _map_headers(fieldnames, overrides)
    missing = [f for f in ("first_name", "last_name", "email")
               if f not in mapping]
    if missing:
        return [], ("Pflichtspalten fehlen: " + ", ".join(missing) +
                    ". Erkannte Spalten: " + ", ".join(filter(None, fieldnames)))
    rows = []
    for line, values in enumerate(it, start=2):
        if line - 1 > MAX_ROWS:
            wb.close()
            return [], f"Mehr als {MAX_ROWS} Zeilen – bitte Datei aufteilen."
        raw = {fieldnames[i]: ("" if v is None else str(v).strip())
               for i, v in enumerate(values or ()) if i < len(fieldnames)}
        if not any(raw.values()):
            continue  # Leerzeilen ueberspringen (Excel-Klassiker)
        # mapping = {feld: quellspalte} – identisch zu parse_csv
        row = {field: raw.get(col, "") for field, col in mapping.items()}
        row["_line"] = line
        rows.append(row)
    wb.close()
    return rows, None


def match_cv_files(zip_bytes, dry_run=True, max_file_mb=10):
    """CV-Dateiberg aus dem Altsystem den Bewerbungen zuordnen.

    Konvention (im UI erklaert): Dateiname beginnt mit der E-Mail-Adresse,
    z. B. 'maria.weber@web.de_Lebenslauf.pdf' oder 'm.weber@gmx.de.pdf'.
    Zuordnung ueber den Blind-Index (E-Mails sind verschluesselt gespeichert);
    Ziel ist die JUENGSTE Bewerbung der Person. dry_run liefert nur den
    Report – importiert wird erst nach Sichtung (wie beim CSV-Import).
    """
    import io as _io
    import os
    import zipfile

    from django.core.files.base import ContentFile
    from django.core.files.storage import default_storage

    from .models import Applicant, Application, ApplicationDocument, email_blind_index

    ALLOWED = {".pdf", ".doc", ".docx", ".jpg", ".jpeg", ".png"}
    report = {"total": 0, "matched": [], "unmatched": [], "errors": [],
              "attached": 0}
    try:
        zf = zipfile.ZipFile(_io.BytesIO(zip_bytes))
    except zipfile.BadZipFile:
        report["errors"].append(("-", "Datei ist kein lesbares ZIP."))
        return report
    for info in zf.infolist():
        if info.is_dir():
            continue
        base = os.path.basename(info.filename)  # Pfad-Traversal neutralisiert
        if not base or base.startswith("."):
            continue
        report["total"] += 1
        stem, ext = os.path.splitext(base)
        if ext.lower() not in ALLOWED:
            report["errors"].append((base, f"Dateityp {ext} nicht erlaubt."))
            continue
        if info.file_size > max_file_mb * 1024 * 1024:
            report["errors"].append((base, f"Größer als {max_file_mb} MB."))
            continue
        email_part = stem.split("_")[0].strip().lower()
        if "@" not in email_part:
            report["unmatched"].append((base, "Kein E-Mail-Präfix im Namen."))
            continue
        applicant = Applicant.objects.filter(
            emailHash=email_blind_index(email_part)).first()
        if applicant is None:
            report["unmatched"].append((base, f"Kein Bewerber: {email_part}"))
            continue
        app = (Application.objects.filter(applicant=applicant)
               .order_by("-createdAt").first())
        if app is None:
            report["unmatched"].append((base, f"Keine Bewerbung: {email_part}"))
            continue
        doc_type = ("CV" if any(k in stem.lower()
                                for k in ("lebenslauf", "cv", "resume"))
                    else "OTHER")
        report["matched"].append((base, email_part, app.jobPosting.title,
                                  doc_type))
        if not dry_run:
            safe = base.replace("..", ".")
            path = default_storage.save(f"application_docs/{safe}",
                                        ContentFile(zf.read(info)))
            ApplicationDocument.objects.create(
                application=app, name=base, file=path, docType=doc_type)
            report["attached"] += 1
    return report
